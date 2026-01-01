#!/usr/bin/env python3

from __future__ import annotations

import os
import re
import sys
import ssl
import json
import time
import socket
import subprocess
import urllib.request
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Dict, Any, List, Optional, Tuple

import ipaddress
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SCRIPT_VERSION = "2.6.4"

# ---------------------------------------------------------------------
# Logging / Rate limiting / Simple cache
# ---------------------------------------------------------------------
# API_CACHE משמש כ-cache בזיכרון כדי לא לחזור על קריאות HTTP זהות לאותו IP/endpoint.
# API_RATE_LIMITS מגדיר השהיות מינימליות בין קריאות כדי לא “להפציץ” APIs ציבוריים.
API_CACHE: Dict[str, Any] = {}

API_RATE_LIMITS = {
    "cloudflare": {"last_call": 0.0, "min_interval": 0.1},
    "cloudflare_rpki": {"last_call": 0.0, "min_interval": 0.1},
    "keycdn": {"last_call": 0.0, "min_interval": 0.5},
    "ripestat": {"last_call": 0.0, "min_interval": 0.2},
    "bgpview": {"last_call": 0.0, "min_interval": 0.5},
    "ipapi": {"last_call": 0.0, "min_interval": 1.5},
}

def log(msg: str) -> None:
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")

def rate_limit(api_name: str) -> None:
    limit = API_RATE_LIMITS.get(api_name)
    if not limit:
        return
    elapsed = time.time() - limit["last_call"]
    if elapsed < limit["min_interval"]:
        time.sleep(limit["min_interval"] - elapsed)
    limit["last_call"] = time.time()

def http_get_json(url: str, headers: Optional[dict] = None, timeout: int = 10) -> Optional[dict]:
    # עטיפה פשוטה ל-GET שמחזירה dict אם הצליח, אחרת None.
    try:
        req = urllib.request.Request(url, headers=headers or {})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode())
    except Exception:
        return None

def is_unknown(v: Any) -> bool:
    # פונקציה אחידה לזיהוי ערכים “לא ידועים”/ריקים.
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.upper() == "UNKNOWN" or s.upper() == "N/A"

# ---------------------------------------------------------------------
# Cloudflare RPKI ROAs dataset (TA in ROA)
# ---------------------------------------------------------------------
# זה בסיס הסיווג של "איזה TA מכסה את ה-prefix של IP".
# אנחנו מורידים את rpki.json של Cloudflare (ROAs dataset), ומכינים מראש רשימת networks
# כדי שנוכל לבצע lookup מהיר: IP -> אילו ROA prefixes מכסים אותו + אילו TAs קשורים אליהם.

def normalize_ta_name(ta: str) -> str:
    # נרמול שם TA כדי לקבל סט קבוע: AFRINIC/APNIC/ARIN/LACNIC/RIPE
    ta_upper = (ta or "").upper()
    if "AFRINIC" in ta_upper:
        return "AFRINIC"
    if "APNIC" in ta_upper:
        return "APNIC"
    if "ARIN" in ta_upper:
        return "ARIN"
    if "LACNIC" in ta_upper:
        return "LACNIC"
    if "RIPE" in ta_upper:
        return "RIPE"
    return (ta or "").upper()

def get_cloudflare_rpki_roas() -> Dict[str, List[dict]]:
    """
    מוריד את dataset של ROAs מ-rpki.cloudflare.com ומבצע pre-parse ל-networks:
      - roa_data: prefix(string) -> list[{ta, asn, maxLength}]
      - cloudflare_rpki_networks: list[(ip_network(prefix), entries)]
    המטרה: לאפשר find_roa_covering_tas_and_prefixes(ip) בצורה מהירה.
    """
    cache_key = "cloudflare_rpki_roas"
    if cache_key in API_CACHE:
        return API_CACHE[cache_key]

    log("  Downloading ROAs from rpki.cloudflare.com...")
    rate_limit("cloudflare_rpki")

    url = "https://rpki.cloudflare.com/rpki.json"
    data = http_get_json(
        url,
        headers={"User-Agent": f"RPKI-Analyzer/{SCRIPT_VERSION}", "Accept": "application/json"},
        timeout=60,
    )

    roa_data: Dict[str, List[dict]] = {}
    parsed_networks: List[Tuple[ipaddress._BaseNetwork, List[dict]]] = []

    if not data or "roas" not in data:
        log("  ERROR: Failed to load Cloudflare RPKI data (no JSON/roas).")
        API_CACHE["cloudflare_rpki_networks"] = []
        API_CACHE[cache_key] = {}
        return {}

    roas = data.get("roas", [])
    log(f"  Loaded {len(roas)} ROAs from Cloudflare RPKI")

    # בונים מיפוי prefix -> entries (עם TA מנורמל)
    for roa in roas:
        prefix = roa.get("prefix", "")
        ta = roa.get("ta", "")
        asn = roa.get("asn", "")
        max_length = roa.get("maxLength", 0)
        if not prefix or not ta:
            continue
        ta_norm = normalize_ta_name(ta)
        roa_data.setdefault(prefix, []).append({"ta": ta_norm, "asn": asn, "maxLength": max_length})

    # pre-parse של prefix strings ל-ip_network כדי לבדוק membership מהר: ip in net
    for prefix, entries in roa_data.items():
        try:
            net = ipaddress.ip_network(prefix, strict=False)
            parsed_networks.append((net, entries))
        except ValueError:
            continue

    API_CACHE["cloudflare_rpki_networks"] = parsed_networks
    API_CACHE[cache_key] = roa_data
    log(f"  Pre-parsed {len(parsed_networks)} ROA networks for fast lookup")
    return roa_data

def find_roa_covering_tas_and_prefixes(ip: str) -> Tuple[List[str], List[str]]:
    """
    “סיווג IP לפי ROA coverage”:
    מקבל IP ומחזיר:
      1) רשימת TAs (ללא כפילויות) שיש ROA prefixes שמכסים את ה-IP
      2) רשימת prefix samples בפורמט: "prefix/len (TA)" להסבר בדוח
    """
    try:
        ip_obj = ipaddress.ip_address(ip)
    except ValueError:
        return [], []

    nets = API_CACHE.get("cloudflare_rpki_networks", [])
    tas = set()
    prefixes = []

    # עוברים על כל ROA networks ומזהים membership של ה-IP
    for net, roas in nets:
        if net.version != ip_obj.version:
            continue
        if ip_obj in net:
            for roa in roas:
                ta = roa.get("ta")
                if ta:
                    tas.add(ta)
                    prefixes.append(f"{net.with_prefixlen} ({ta})")

    # de-dup עבור prefix samples
    seen = set()
    dedup = []
    for p in prefixes:
        if p not in seen:
            seen.add(p)
            dedup.append(p)

    return sorted(tas), dedup

# ---------------------------------------------------------------------
# TA Origin inference
# ---------------------------------------------------------------------
# "TA Origin" = ה-TA של *התוכן* (RPKI objects) ולא של ה-IP.
# אנחנו מסיקים אותו מרמזים ב-SIA/AIA של certificates / ROA-embedded cert,
# או fallback לפי path, או RIR allocation, או union יחיד של ROA coverage.

def get_ta_origin_from_sia(sia_repo: str) -> Optional[str]:
    # חיפוש תבניות TA בתוך SIA/AIA URI (למשל rpki.ripe.net וכו').
    if not sia_repo:
        return None
    s = sia_repo.lower()
    ta_patterns = {
        "AFRINIC": ["afrinic", "rpki.afrinic.net"],
        "APNIC": ["apnic", "rpki.apnic.net"],
        "ARIN": ["arin", "rpki.arin.net"],
        "LACNIC": ["lacnic", "rpki.lacnic.net", "repository.lacnic.net", "rrdp.lacnic.net"],
        "RIPE": ["ripe", "rpki.ripe.net", "rrdp.ripe.net", "paas.rpki.ripe.net"],
    }
    for ta, pats in ta_patterns.items():
        if any(p in s for p in pats):
            return ta
    return None

def get_ta_origin_from_path(file_path: str) -> Optional[str]:
    # fallback חלש: אם הנתיב כולל מחרוזות שמרמזות על TA.
    if not file_path:
        return None
    p = file_path.lower()
    ta_patterns = {
        "AFRINIC": ["afrinic", "/afrinic/", "rpki.afrinic"],
        "APNIC": ["apnic", "/apnic/", "rpki.apnic"],
        "ARIN": ["arin", "/arin/", "rpki.arin"],
        "LACNIC": ["lacnic", "/lacnic/", "rpki.lacnic"],
        "RIPE": ["ripe", "/ripe/", "rpki.ripe", "rrdp.ripe", "paas.rpki.ripe"],
    }
    for ta, pats in ta_patterns.items():
        if any(x in p for x in pats):
            return ta
    return None

# ---------------------------------------------------------------------
# IP info APIs
# ---------------------------------------------------------------------
# המטרה פה אינה לקבוע TA של התוכן, אלא:
#   - לזהות ASN/Org/RIR של ה-IP
#   - לזהות אם ה-IP נראה כ-hosting/CDN (ip-api “hosting”)
# זה משמש אחר כך להחלטה האם CROSS_TA הוא "אמיתי" או רק Delivery mismatch דרך CDN.

def get_cloudflare_radar_info(ip: str) -> Optional[dict]:
    cache_key = f"cloudflare_{ip}"
    if cache_key in API_CACHE:
        return API_CACHE[cache_key]

    rate_limit("cloudflare")
    url = f"https://api.cloudflare.com/client/v4/radar/entities/asns/ip?ip={ip}"
    data = http_get_json(
        url,
        headers={"User-Agent": f"RPKI-Analyzer/{SCRIPT_VERSION}", "Accept": "application/json"},
        timeout=10,
    )
    if not data or not data.get("success") or "result" not in data:
        return None

    asn_data = (data["result"] or {}).get("asn", {}) or {}
    asn = asn_data.get("asn")
    org_name = asn_data.get("name", "") or ""
    org_full = asn_data.get("orgName", "") or ""
    country = asn_data.get("country", "") or ""
    source = asn_data.get("source", "") or ""

    # Cloudflare נותן "source" שמרמז על ה-RIR; נתרגם אותו לסט קבוע.
    rir = None
    su = source.upper()
    if "AFRINIC" in su:
        rir = "AFRINIC"
    elif "APNIC" in su:
        rir = "APNIC"
    elif "ARIN" in su:
        rir = "ARIN"
    elif "LACNIC" in su:
        rir = "LACNIC"
    elif "RIPE" in su:
        rir = "RIPE"

    res = {
        "asn": asn,
        "org_name": org_name,
        "org_full": org_full or org_name,
        "country": country,
        "rir": rir,
        "source": "cloudflare",
    }
    API_CACHE[cache_key] = res
    return res

def get_keycdn_info(ip: str) -> Optional[dict]:
    # fallback עבור ISP/ASN כאשר אין org_name טוב.
    cache_key = f"keycdn_{ip}"
    if cache_key in API_CACHE:
        return API_CACHE[cache_key]

    rate_limit("keycdn")
    url = f"https://tools.keycdn.com/geo.json?host={ip}"
    data = http_get_json(
        url,
        headers={"User-Agent": "keycdn-tools:https://rpki-analyzer.app", "Accept": "application/json"},
        timeout=10,
    )
    if not data or data.get("status") != "success":
        return None

    geo = (data.get("data") or {}).get("geo", {}) or {}
    res = {
        "asn": geo.get("asn"),
        "isp": geo.get("isp", "") or "",
        "country": geo.get("country_code", "") or "",
        "source": "keycdn",
    }
    API_CACHE[cache_key] = res
    return res

def get_ripestat_rir_allocation(ip: str) -> Optional[str]:
    # מקור אמין ל-RIR allocation של כתובת/משאב.
    cache_key = f"ripestat_rir_{ip}"
    if cache_key in API_CACHE:
        return API_CACHE[cache_key]

    rate_limit("ripestat")
    url = f"https://stat.ripe.net/data/rir/data.json?resource={ip}"
    data = http_get_json(
        url,
        headers={"User-Agent": f"RPKI-Analyzer/{SCRIPT_VERSION}", "Accept": "application/json"},
        timeout=10,
    )
    if not data or data.get("status") != "ok":
        return None

    rirs = (data.get("data") or {}).get("rirs", []) or []
    if not rirs:
        return None

    rir_name = ((rirs[0] or {}).get("rir", "") or "").upper()
    if "AFRINIC" in rir_name:
        API_CACHE[cache_key] = "AFRINIC"
    elif "APNIC" in rir_name:
        API_CACHE[cache_key] = "APNIC"
    elif "ARIN" in rir_name:
        API_CACHE[cache_key] = "ARIN"
    elif "LACNIC" in rir_name:
        API_CACHE[cache_key] = "LACNIC"
    elif "RIPE" in rir_name or "RIPENCC" in rir_name:
        API_CACHE[cache_key] = "RIPE"
    return API_CACHE.get(cache_key)

def get_bgpview_info(ip: str) -> Optional[dict]:
    # fallback ל-ASN/prefix/RIR info דרך BGPView.
    cache_key = f"bgpview_{ip}"
    if cache_key in API_CACHE:
        return API_CACHE[cache_key]

    rate_limit("bgpview")
    url = f"https://api.bgpview.io/ip/{ip}"
    data = http_get_json(url, headers={"User-Agent": f"RPKI-Analyzer/{SCRIPT_VERSION}"}, timeout=10)
    if not data or data.get("status") != "ok":
        return None

    ip_data = data.get("data", {}) or {}
    prefixes = ip_data.get("prefixes", []) or []
    rir_alloc = ip_data.get("rir_allocation", {}) or {}

    asn = None
    asn_name = ""
    asn_desc = ""
    prefix = ""

    if prefixes:
        p0 = prefixes[0] or {}
        asn_info = p0.get("asn", {}) or {}
        asn = asn_info.get("asn")
        asn_name = asn_info.get("name", "") or ""
        asn_desc = asn_info.get("description", "") or ""
        prefix = p0.get("prefix", "") or ""

    rir = None
    rir_name = (rir_alloc.get("rir_name", "") or "").upper()
    if "AFRINIC" in rir_name:
        rir = "AFRINIC"
    elif "APNIC" in rir_name:
        rir = "APNIC"
    elif "ARIN" in rir_name:
        rir = "ARIN"
    elif "LACNIC" in rir_name:
        rir = "LACNIC"
    elif "RIPE" in rir_name or "RIPENCC" in rir_name:
        rir = "RIPE"

    res = {
        "asn": asn,
        "asn_name": asn_name,
        "asn_desc": asn_desc,
        "prefix": prefix,
        "rir": rir,
        "source": "bgpview",
    }
    API_CACHE[cache_key] = res
    return res

def get_ipapi_info(ip: str) -> Optional[dict]:
    # ip-api נותן דגל hosting, שימושי לזיהוי CDN/Datacenter delivery.
    cache_key = f"ipapi_{ip}"
    if cache_key in API_CACHE:
        return API_CACHE[cache_key]

    rate_limit("ipapi")
    url = f"http://ip-api.com/json/{ip}?fields=status,org,as,isp,hosting"
    data = http_get_json(url, headers={"User-Agent": f"RPKI-Analyzer/{SCRIPT_VERSION}"}, timeout=6)
    if not data or data.get("status") != "success":
        return None

    org = data.get("org", "") or ""
    isp = data.get("isp", "") or ""
    as_info = data.get("as", "") or ""
    hosting = bool(data.get("hosting", False))

    asn = None
    m = re.match(r"AS(\d+)", as_info)
    if m:
        asn = int(m.group(1))

    res = {
        "org": org,
        "isp": isp,
        "asn": asn,
        "as_info": as_info,
        "is_hosting": hosting,
        "source": "ip-api",
    }
    API_CACHE[cache_key] = res
    return res

def get_comprehensive_ip_info(ip: str) -> Dict[str, Any]:
    """
    מאגד מידע על IP מכמה APIs לפי סדר עדיפות:
      - Cloudflare Radar: ASN + org + RIR (טוב ומהיר)
      - KeyCDN Geo: fallback ל-org/ISP
      - RIPEstat: RIR allocation (אם חסר)
      - BGPView: fallback ל-ASN/RIR/prefix
      - ip-api: hosting flag + hosting provider (להחלטת CDN/hosting)
    הפלט משמש ל:
      - "hosting_org/rir/asn" בדוח
      - סימון is_cdn כאשר delivery נראה כ-CDN/hosting
    """
    cache_key = f"comprehensive_{ip}"
    if cache_key in API_CACHE:
        return API_CACHE[cache_key]

    result = {
        "asn": None,
        "org_name": None,
        "rir": None,
        "hosting_provider": None,
        "is_hosting": False,
        "source": None,
    }

    cf = get_cloudflare_radar_info(ip)
    if cf:
        result["asn"] = cf.get("asn")
        result["org_name"] = cf.get("org_full") or cf.get("org_name")
        result["rir"] = cf.get("rir")
        result["source"] = "cloudflare"

    if not result["org_name"]:
        kc = get_keycdn_info(ip)
        if kc:
            if not result["asn"]:
                result["asn"] = kc.get("asn")
            result["org_name"] = kc.get("isp")
            result["source"] = result["source"] or "keycdn"

    if not result["rir"]:
        rir = get_ripestat_rir_allocation(ip)
        if rir:
            result["rir"] = rir

    if not result["asn"] or not result["rir"]:
        bgp = get_bgpview_info(ip)
        if bgp:
            if not result["asn"]:
                result["asn"] = bgp.get("asn")
            if not result["rir"]:
                result["rir"] = bgp.get("rir")
            if not result["org_name"]:
                result["org_name"] = bgp.get("asn_name") or bgp.get("asn_desc")
            result["source"] = result["source"] or "bgpview"

    ipa = get_ipapi_info(ip)
    if ipa:
        if not result["asn"]:
            result["asn"] = ipa.get("asn")
        result["is_hosting"] = bool(ipa.get("is_hosting", False))
        if result["is_hosting"]:
            result["hosting_provider"] = ipa.get("org") or ipa.get("isp")
        if not result["org_name"]:
            result["org_name"] = ipa.get("org") or ipa.get("isp")
        result["source"] = result["source"] or "ip-api"

    API_CACHE[cache_key] = result
    return result

# ---------------------------------------------------------------------
# Networking helpers
# ---------------------------------------------------------------------
# resolve_domain() היא השלב שמייצר את "כתובות ה-delivery" של repository domain.
# בהמשך, כתובות אלו מושוות מול ROA coverage כדי לקבוע SAME_TA / CROSS_TA וכו'.

def is_ipv4(s: str) -> bool:
    try:
        return ipaddress.ip_address(s).version == 4
    except ValueError:
        return False

def is_ipv6(s: str) -> bool:
    try:
        return ipaddress.ip_address(s).version == 6
    except ValueError:
        return False

def resolve_domain(domain: str) -> Dict[str, List[str]]:
    """
    פותר domain לכתובות IPv4/IPv6 (delivery IPs):
      1) socket.getaddrinfo (מהיר)
      2) fallback ל-dig A/AAAA אם חסר
    הפלט:
      {"v4": [...], "v6": [...]}
    """
    out = {"v4": [], "v6": []}
    if not domain or domain == "unknown_repository":
        return out

    domain = domain.strip()
    if "." not in domain or "/" in domain or len(domain) > 200:
        return out

    try:
        socket.setdefaulttimeout(10)
        res = socket.getaddrinfo(domain, None)
        for r in res:
            ip = r[4][0]
            if is_ipv4(ip) and ip not in out["v4"]:
                out["v4"].append(ip)
            elif is_ipv6(ip) and ip not in out["v6"]:
                out["v6"].append(ip)
    except Exception:
        pass

    if not out["v4"]:
        try:
            p = subprocess.run(["dig", "+short", "+time=8", "+tries=2", "A", domain],
                               capture_output=True, text=True, timeout=20)
            if p.returncode == 0 and p.stdout:
                for line in p.stdout.splitlines():
                    line = line.strip()
                    if is_ipv4(line) and line not in out["v4"]:
                        out["v4"].append(line)
        except Exception:
            pass

    if not out["v6"]:
        try:
            p = subprocess.run(["dig", "+short", "+time=8", "+tries=2", "AAAA", domain],
                               capture_output=True, text=True, timeout=20)
            if p.returncode == 0 and p.stdout:
                for line in p.stdout.splitlines():
                    line = line.strip()
                    if is_ipv6(line) and line not in out["v6"]:
                        out["v6"].append(line)
        except Exception:
            pass

    return out

def check_http_headers_for_cdn(domain: str) -> Optional[str]:
    """
    זיהוי CDN על בסיס headers ב-HTTPS response.
    זה fallback משלים לזיהוי CDN/hosting (בנוסף ל-ip-api hosting flag).
    """
    cdn_headers = {
        "cf-ray": "Cloudflare",
        "cf-cache-status": "Cloudflare",
        "x-amz-cf-id": "AWS CloudFront",
        "x-amz-cf-pop": "AWS CloudFront",
        "x-served-by": "Fastly",
        "x-fastly-request-id": "Fastly",
        "x-akamai-transformed": "Akamai",
        "x-edge-location": "Akamai",
        "x-azure-ref": "Microsoft Azure CDN",
        "x-msedge-ref": "Microsoft Azure CDN",
        "x-sucuri-id": "Sucuri",
        "x-sucuri-cache": "Sucuri",
        "x-iinfo": "Incapsula",
        "x-hw": "Huawei Cloud CDN",
        "eagleid": "Alibaba Cloud CDN",
    }
    try:
        req = urllib.request.Request(
            f"https://{domain}/",
            headers={"User-Agent": f"Mozilla/5.0 RPKI-Analyzer/{SCRIPT_VERSION}"},
        )
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        with urllib.request.urlopen(req, timeout=5, context=ctx) as resp:
            headers = {k.lower(): v for k, v in resp.headers.items()}
        for h, name in cdn_headers.items():
            if h in headers:
                return name
        if "x-cdn" in headers:
            return headers["x-cdn"]
    except Exception:
        return None
    return None

# ---------------------------------------------------------------------
# Path parsing (repo domain)
# ---------------------------------------------------------------------
# extract_domain_from_path() ממפה קבצים בדיסק ל-repository "domain".
# זה מאפשר לקבץ קבצים באותו repo תחת מפתח אחד במילון repositories.

def extract_domain_from_path(path: str) -> str:
    parts = path.split(os.sep)

    for part in parts:
        if part in [".", "..", "cache", "repository", "rrdp", "ta", "repo", ""]:
            continue
        if part.endswith((".cer", ".roa", ".mft", ".crl", ".xml")):
            continue
        if part.startswith("."):
            continue
        if len(part) >= 20 and re.match(r"^[A-Za-z0-9_-]+$", part) and "." not in part:
            continue
        if re.match(r"^[a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12}$", part, re.I):
            continue
        if "." in part and re.match(r"^[A-Za-z0-9][A-Za-z0-9.-]+\.[A-Za-z]{2,}$", part):
            return part

    for i, part in enumerate(parts):
        if part == "cache" and i + 1 < len(parts):
            nxt = parts[i + 1]
            if "." in nxt and not nxt.startswith("."):
                return nxt

    return "unknown_repository"

# ---------------------------------------------------------------------
# OpenSSL parsing
# ---------------------------------------------------------------------
# parse_certificate() מחלץ SIA/AIA מה-cert. זה קריטי ל-"TA Origin" של התוכן.
# run_openssl_cms_print + parse_roa_embedded_sia_and_aia נותנים hints גם מתוך ROA.

def run_openssl_text(cer_path: str) -> Optional[str]:
    try:
        p = subprocess.run(
            ["openssl", "x509", "-inform", "DER", "-in", cer_path, "-noout", "-text"],
            capture_output=True,
            text=True,
            timeout=12,
        )
        return p.stdout if p.returncode == 0 and p.stdout else None
    except Exception:
        return None

def parse_sia_block(cert_text: str) -> Tuple[str, str]:
    m = re.search(
        r"Subject Information Access:\s*\n(.*?)(?=\n\s*(?:Authority Information Access:|sbgp-|X509v3|Signature Algorithm:|$))",
        cert_text,
        re.DOTALL | re.IGNORECASE,
    )
    if not m:
        return "", ""

    block = m.group(1)
    repo = ""
    mft = ""

    mr = re.search(r"CA\s*Repository\s*-\s*URI:(\S+)", block, re.IGNORECASE)
    if mr:
        repo = mr.group(1).strip()

    mm = re.search(r"RPKI\s*Manifest\s*-\s*URI:(\S+)", block, re.IGNORECASE)
    if mm:
        mft = mm.group(1).strip()

    return repo, mft

def parse_certificate(cer_path: str) -> Optional[Dict[str, str]]:
    cert_text = run_openssl_text(cer_path)
    if not cert_text:
        return None

    info: Dict[str, str] = {
        "path": cer_path,
        "filename": os.path.basename(cer_path),
        "subject": "",
        "issuer": "",
        "not_before": "",
        "not_after": "",
        "sia_repo": "",
        "sia_mft": "",
        "aia": "",
        "ip_resources": "",
        "as_resources": "",
    }

    m = re.search(r"Subject:\s*(.+)", cert_text)
    if m:
        info["subject"] = m.group(1).strip()

    m = re.search(r"Issuer:\s*(.+)", cert_text)
    if m:
        info["issuer"] = m.group(1).strip()

    m = re.search(r"Not Before:\s*(.+)", cert_text)
    if m:
        info["not_before"] = m.group(1).strip()

    m = re.search(r"Not After\s*:\s*(.+)", cert_text)
    if m:
        info["not_after"] = m.group(1).strip()

    # SIA: CA Repository URI (רמז חזק ל-TA של התוכן)
    # SIA: Manifest URI (לא קריטי ל-TA, אבל נשמר בדוח)
    repo, mft = parse_sia_block(cert_text)
    info["sia_repo"] = repo
    info["sia_mft"] = mft

    # AIA: CA Issuers URI (רמז חזק נוסף ל-TA של התוכן)
    aia = re.search(
        r"Authority Information Access:\s*\n(.*?)(?=\n\s*(?:sbgp-|X509v3|Signature Algorithm:|$))",
        cert_text,
        re.DOTALL | re.IGNORECASE,
    )
    if aia:
        block = aia.group(1)
        ca = re.search(r"CA\s*Issuers\s*-\s*URI:(\S+)", block, re.IGNORECASE)
        if ca:
            info["aia"] = ca.group(1).strip()

    return info

def run_openssl_cms_print(roa_path: str) -> Optional[str]:
    try:
        p = subprocess.run(
            ["openssl", "cms", "-inform", "DER", "-in", roa_path, "-noout", "-cmsout", "-print"],
            capture_output=True,
            text=True,
            timeout=15,
        )
        return p.stdout if p.returncode == 0 and p.stdout else None
    except Exception:
        return None

def parse_roa_embedded_sia_and_aia(cms_text: str) -> Tuple[str, str]:
    # ROA הוא CMS שכולל cert/מקטעים שבהם ניתן למצוא AIA/SIA.
    # זה מאפשר להסיק TA origin גם כשאין cert sample זמין.
    sia_repo = ""
    aia_uri = ""

    m = re.search(r"Authority Information Access.*?CA\s*Issuers\s*-\s*URI:(\S+)", cms_text, re.I | re.S)
    if m:
        aia_uri = m.group(1).strip()

    m = re.search(r"Subject Information Access.*?CA\s*Repository\s*-\s*URI:(\S+)", cms_text, re.I | re.S)
    if m:
        sia_repo = m.group(1).strip()

    return sia_repo, aia_uri

# ---------------------------------------------------------------------
# Expected TA chooser + helper to infer TA source even if TA exists
# ---------------------------------------------------------------------
# choose_expected_ta_with_source:
#   מייצר "Expected TA" (TA של התוכן) + מקור ההסקה (CERT_SIA/CERT_AIA/ROA_*/HOSTING_RIR/ROA_UNION/UNKNOWN).
#
# למה זה חשוב ל"סיווג כתובות"?
#   כי אחר כך נשווה Expected TA (תוכן) מול union של ROA coverage על delivery IPs.
#   אם יש mismatch נקבל CROSS_TA, אבל נרצה להוריד ל-DELIVERY_TA_MISMATCH אם זה CDN.

def choose_expected_ta_with_source(
    current_ta: str,
    cert_sia_repo: str,
    cert_aia_uri: str,
    roa_sia_repo: str,
    roa_aia_uri: str,
    hosting_rir: str,
    roa_union_tas: List[str],
) -> Tuple[str, str]:
    """
    Returns (expected_ta, source) in this priority:
      EXISTING -> CERT SIA -> CERT AIA -> ROA AIA -> ROA SIA -> Hosting RIR -> single-TA ROA union -> ("", "UNKNOWN")

    פירוט:
      - CERT_SIA / CERT_AIA / ROA_SIA / ROA_AIA = “content-based hints” (אמין יותר)
      - HOSTING_RIR / ROA_UNION = fallback חלש יותר
    """
    if not is_unknown(current_ta):
        return current_ta, "EXISTING"

    cand = get_ta_origin_from_sia(cert_sia_repo)
    if cand:
        return cand, "CERT_SIA"

    cand = get_ta_origin_from_sia(cert_aia_uri)
    if cand:
        return cand, "CERT_AIA"

    cand = get_ta_origin_from_sia(roa_aia_uri)
    if cand:
        return cand, "ROA_AIA"

    cand = get_ta_origin_from_sia(roa_sia_repo)
    if cand:
        return cand, "ROA_SIA"

    if not is_unknown(hosting_rir):
        return str(hosting_rir).strip().upper(), "HOSTING_RIR"

    if len(roa_union_tas) == 1:
        return roa_union_tas[0], "ROA_UNION"

    return "", "UNKNOWN"

def infer_source_for_existing_ta(
    current_ta: str,
    cert_sia_repo: str,
    cert_aia_uri: str,
    roa_sia_repo: str,
    roa_aia_uri: str,
    hosting_rir: str,
    roa_union_tas: List[str],
) -> str:
    """
    אם TA כבר ידוע (מ-Phase 2.5 למשל), מנסים “לשחזר” מאיפה הוא הגיע בפועל.
    למה?
      - בהמשך יש תנאי content_based שמשפיע על הורדה מ-CROSS_TA ל-DELIVERY_TA_MISMATCH.
      - אם נשאיר מקור EXISITING למרות שזה בא מ-SIA/AIA, נפספס את ההורדה ונקבל false positives של CROSS_TA.
    """
    if is_unknown(current_ta):
        return "UNKNOWN"

    ta = str(current_ta).strip().upper()

    cand = get_ta_origin_from_sia(cert_sia_repo)
    if cand and cand == ta:
        return "CERT_SIA"

    cand = get_ta_origin_from_sia(cert_aia_uri)
    if cand and cand == ta:
        return "CERT_AIA"

    cand = get_ta_origin_from_sia(roa_aia_uri)
    if cand and cand == ta:
        return "ROA_AIA"

    cand = get_ta_origin_from_sia(roa_sia_repo)
    if cand and cand == ta:
        return "ROA_SIA"

    if not is_unknown(hosting_rir) and str(hosting_rir).strip().upper() == ta:
        return "HOSTING_RIR"

    if len(roa_union_tas) == 1 and roa_union_tas[0] == ta:
        return "ROA_UNION"

    return "EXISTING"

# ---------------------------------------------------------------------
# ROA coverage status
# ---------------------------------------------------------------------
# compute_roa_coverage_status הוא "סיווג הכתובות" בפועל:
#   Input:
#     - expected_ta: ה-TA של התוכן (מה-SIA/AIA וכו')
#     - ips: delivery IPs שהדומיין resolve אליהם (v4+v6)
#   Output:
#     - union של TAs שמכסים את ה-IPs לפי ROAs
#     - סטטוס:
#         NO_ROA / UNKNOWN_EXPECTED_TA / SAME_TA / CROSS_TA
#     - ta_lost (דגל CROSS_TA אמיתי לפני הורדת CDN mismatch)

def compute_roa_coverage_status(expected_ta: str, ips: List[str], max_ips_to_check: int = 3) -> Dict[str, Any]:
    # מגבילים למקסימום 3 IPs כדי לשמור על ביצועים (DNS יכול להחזיר רשימות ארוכות).
    ips_to_check = [x for x in ips if x][:max_ips_to_check]
    union_tas = set()
    prefixes_sample: List[str] = []

    # לכל IP: מביאים אילו TAs מכסים אותו לפי ROA prefixes, ומבצעים union.
    for ip in ips_to_check:
        tas, prefixes = find_roa_covering_tas_and_prefixes(ip)
        union_tas.update(tas)
        for p in prefixes[:10]:
            if p not in prefixes_sample:
                prefixes_sample.append(p)

    roa_covering_tas = sorted(union_tas)

    # 1) אין ROA שמכסה את ה-delivery IPs -> NO_ROA
    if not roa_covering_tas:
        return {
            "roa_covering_tas": [],
            "roa_covering_prefixes": prefixes_sample[:20],
            "roa_coverage_status": "NO_ROA",
            "ta_lost": False,
            "ta_lost_reason": "",
        }

    # 2) אם אין expected TA, לא מסמנים CROSS_TA (אין בסיס להשוואה)
    if is_unknown(expected_ta):
        return {
            "roa_covering_tas": roa_covering_tas,
            "roa_covering_prefixes": prefixes_sample[:20],
            "roa_coverage_status": "UNKNOWN_EXPECTED_TA",
            "ta_lost": False,
            "ta_lost_reason": "Expected TA unknown; not marking CROSS_TA",
        }

    # 3) expected TA נמצא ב-union -> SAME_TA (תקין)
    if expected_ta in roa_covering_tas:
        return {
            "roa_covering_tas": roa_covering_tas,
            "roa_covering_prefixes": prefixes_sample[:20],
            "roa_coverage_status": "SAME_TA",
            "ta_lost": False,
            "ta_lost_reason": "",
        }

    # 4) expected TA לא נמצא ב-union -> CROSS_TA (חשד לתלות cross-TA/TA_LOST)
    return {
        "roa_covering_tas": roa_covering_tas,
        "roa_covering_prefixes": prefixes_sample[:20],
        "roa_coverage_status": "CROSS_TA",
        "ta_lost": True,
        "ta_lost_reason": f"CROSS_TA_ROA_COVERAGE ({', '.join(roa_covering_tas)})",
    }

# ---------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------
# נקודת המפתח של "סיווג הכתובות" נמצאת ב-Phase 3:
#   (1) resolve_domain -> delivery IPs
#   (2) find_roa_covering... / compute_roa_coverage_status -> TA של delivery
#   (3) choose_expected_ta_with_source -> TA של התוכן
#   (4) השוואה: SAME_TA / CROSS_TA
#   (5) אם CROSS_TA אבל זה CDN + expected TA הגיע מ-content hints => הורדה ל-DELIVERY_TA_MISMATCH

def analyze_rpki_data(cache_dir: str, output_dir: str) -> None:
    cache_path = Path(cache_dir)
    out_path = Path(output_dir)

    if not cache_path.exists():
        print(f"ERROR: Cache directory does not exist: {cache_dir}")
        sys.exit(1)

    out_path.mkdir(parents=True, exist_ok=True)

    log("=" * 60)
    log(f"RPKI Repository Analyzer v{SCRIPT_VERSION}")
    log("=" * 60)

    log("Phase 0: Loading ROA data from Cloudflare RPKI...")
    _ = get_cloudflare_rpki_roas()

    log("Phase 1: Collecting files...")
    all_cer_files = list(cache_path.rglob("*.cer"))
    all_roa_files = list(cache_path.rglob("*.roa"))
    all_mft_files = list(cache_path.rglob("*.mft"))
    all_crl_files = list(cache_path.rglob("*.crl"))

    log(f"  Found {len(all_cer_files)} .cer files")
    log(f"  Found {len(all_roa_files)} .roa files")
    log(f"  Found {len(all_mft_files)} .mft files")
    log(f"  Found {len(all_crl_files)} .crl files")

    log("Phase 2: Identifying repositories...")
    repositories = defaultdict(
        lambda: {
            "domain": "",
            "ta_origin": "",
            "ta_origin_source": "",
            "sample_sia": "",
            "sample_paths": [],
            "sample_roa_paths": [],
            "ips_v4": [],
            "ips_v6": [],
            "primary_ip": "",
            "hosting_rir": "",
            "hosting_asn": None,
            "hosting_org": "",
            "is_cdn": False,
            "cdn_provider": "",
            "roa_covering_tas": "",
            "roa_covering_prefixes": "",
            "roa_coverage_status": "",
            "ta_lost": False,
            "ta_lost_reason": "",
            "delivery_ta_mismatch": False,
            "delivery_ta_mismatch_reason": "",
            "cer_count": 0,
            "roa_count": 0,
            "mft_count": 0,
            "crl_count": 0,
        }
    )

    # קיבוץ הקבצים לפי domain repository (מחלוץ מתוך path).
    for cer_file in all_cer_files:
        rel = str(cer_file.relative_to(cache_path))
        dom = extract_domain_from_path(rel)
        repositories[dom]["domain"] = dom
        repositories[dom]["cer_count"] += 1
        if len(repositories[dom]["sample_paths"]) < 3:
            repositories[dom]["sample_paths"].append(str(cer_file))

    for roa_file in all_roa_files:
        rel = str(roa_file.relative_to(cache_path))
        dom = extract_domain_from_path(rel)
        repositories[dom]["domain"] = dom
        repositories[dom]["roa_count"] += 1
        if len(repositories[dom]["sample_roa_paths"]) < 2:
            repositories[dom]["sample_roa_paths"].append(str(roa_file))

    for mft_file in all_mft_files:
        rel = str(mft_file.relative_to(cache_path))
        dom = extract_domain_from_path(rel)
        repositories[dom]["mft_count"] += 1

    for crl_file in all_crl_files:
        rel = str(crl_file.relative_to(cache_path))
        dom = extract_domain_from_path(rel)
        repositories[dom]["crl_count"] += 1

    log(f"  Found {len(repositories)} unique repositories")

    # repo_hints שומר רמזים "תוכניים" (SIA/AIA) מתוך certs/ROAs עבור כל domain.
    repo_hints: Dict[str, Dict[str, str]] = defaultdict(lambda: {
        "cert_sia": "", "cert_aia": "", "roa_sia": "", "roa_aia": ""
    })

    log("Phase 2.5: Derive TA Origin + sample SIA from .cer or ROA-embedded cert...")
    for dom, repo in repositories.items():
        # ניסיון ראשון: לקחת cert sample ולחלץ ממנו SIA/AIA כדי להסיק TA Origin של התוכן.
        if repo["sample_paths"]:
            cert_path = repo["sample_paths"][0]
            ci = parse_certificate(cert_path)
            if ci:
                repo_hints[dom]["cert_sia"] = ci.get("sia_repo", "") or ""
                repo_hints[dom]["cert_aia"] = ci.get("aia", "") or ""

                # sample_sia נשמר לדוח כדי להראות “למה חשבנו שזה TA X”
                if ci.get("sia_repo"):
                    repo["sample_sia"] = ci["sia_repo"]
                    ta = get_ta_origin_from_sia(ci["sia_repo"])
                    if ta:
                        # כאן קובעים TA Origin של התוכן לפי SIA (רמז content-based חזק)
                        repo["ta_origin"] = ta  # source fixed later deterministically

                # fallback חלש יותר: לפי path
                if is_unknown(repo["ta_origin"]):
                    ta = get_ta_origin_from_path(cert_path)
                    if ta:
                        repo["ta_origin"] = ta

        # אם עדיין אין TA, ננסה לחלץ hints מתוך ROA-embedded cert (CMS print)
        if is_unknown(repo["ta_origin"]) and repo["sample_roa_paths"]:
            roa_path = repo["sample_roa_paths"][0]
            cms_text = run_openssl_cms_print(roa_path)
            if cms_text:
                sia_repo, aia_uri = parse_roa_embedded_sia_and_aia(cms_text)
                repo_hints[dom]["roa_sia"] = sia_repo or ""
                repo_hints[dom]["roa_aia"] = aia_uri or ""

                if is_unknown(repo["sample_sia"]) and sia_repo:
                    repo["sample_sia"] = sia_repo

                ta = get_ta_origin_from_sia(aia_uri) or get_ta_origin_from_sia(sia_repo)
                if ta:
                    repo["ta_origin"] = ta  # source fixed later deterministically

    log("Phase 3: Domain resolution + hosting info + ROA coverage analysis...")
    resolved_count = 0
    failed_count = 0
    cross_ta_count = 0
    delivery_mismatch_count = 0

    for i, (dom, repo) in enumerate(repositories.items(), 1):
        if i % 10 == 0:
            log(f"  Progress: {i}/{len(repositories)}")

        # (A) Resolve domain -> delivery IPs
        r = resolve_domain(dom)
        v4 = r["v4"]
        v6 = r["v6"]
        repo["ips_v4"] = v4
        repo["ips_v6"] = v6

        # primary_ip = כתובת “מייצגת” לצורך ASN/RIR/hosting info
        primary = v4[0] if v4 else (v6[0] if v6 else "")
        repo["primary_ip"] = primary

        if not primary:
            failed_count += 1
            continue
        resolved_count += 1

        # (B) פרופיל IP: ASN/Org/RIR + האם זה hosting/CDN
        ip_info = get_comprehensive_ip_info(primary)
        repo["hosting_asn"] = ip_info.get("asn")
        repo["hosting_org"] = ip_info.get("org_name") or ""
        repo["hosting_rir"] = ip_info.get("rir") or ""

        # (C) union של TAs שמכסים את delivery IPs לפי ROAs (עד 3 IPs)
        ips_for_roa = (v4 + v6)
        union_tas = set()
        for ip in ips_for_roa[:3]:
            tas, _ = find_roa_covering_tas_and_prefixes(ip)
            union_tas.update(tas)
        union_tas_list = sorted(union_tas)

        # (D) expected TA = TA של התוכן (לא של ה-IP) + מקור ההסקה
        exp_ta, exp_src = choose_expected_ta_with_source(
            repo.get("ta_origin", ""),
            repo_hints[dom].get("cert_sia", ""),
            repo_hints[dom].get("cert_aia", ""),
            repo_hints[dom].get("roa_sia", ""),
            repo_hints[dom].get("roa_aia", ""),
            repo.get("hosting_rir", ""),
            union_tas_list,
        )

        # NEW FIX:
        # אם היה TA "קיים" (EXISTING), ננסה לסווג מאיפה הוא הגיע בפועל.
        # זה חשוב כדי לזהות content_based נכון ולהוריד CROSS_TA ל-DELIVERY_TA_MISMATCH כשצריך.
        if exp_src == "EXISTING":
            exp_src = infer_source_for_existing_ta(
                exp_ta,
                repo_hints[dom].get("cert_sia", ""),
                repo_hints[dom].get("cert_aia", ""),
                repo_hints[dom].get("roa_sia", ""),
                repo_hints[dom].get("roa_aia", ""),
                repo.get("hosting_rir", ""),
                union_tas_list,
            )

        repo["ta_origin"] = exp_ta
        repo["ta_origin_source"] = exp_src

        # (E) זיהוי CDN/Hosting:
        # 1) ip-api hosting flag (אם יש provider)
        # 2) fallback: בדיקת HTTP headers שמרמזים CDN
        if ip_info.get("is_hosting") and ip_info.get("hosting_provider"):
            repo["is_cdn"] = True
            repo["cdn_provider"] = ip_info["hosting_provider"]

        if not repo["is_cdn"]:
            cdn = check_http_headers_for_cdn(dom)
            if cdn:
                repo["is_cdn"] = True
                repo["cdn_provider"] = cdn

        # (F) “סיווג הכתובות”:
        # משווים expected TA (תוכן) מול union של ROA coverage על delivery IPs:
        #   - NO_ROA / SAME_TA / CROSS_TA / ...
        cov = compute_roa_coverage_status(repo["ta_origin"], ips_for_roa, max_ips_to_check=3)

        repo["roa_covering_tas"] = ", ".join(cov["roa_covering_tas"])
        repo["roa_covering_prefixes"] = ", ".join(cov["roa_covering_prefixes"])
        repo["roa_coverage_status"] = cov["roa_coverage_status"]
        repo["ta_lost"] = cov["ta_lost"]
        repo["ta_lost_reason"] = cov["ta_lost_reason"]

        # content_based = האם TA של התוכן הגיע מרמזים תוכניים (SIA/AIA),
        # כלומר אמין יותר. זה תנאי להורדה ל-DELIVERY_TA_MISMATCH.
        content_based = repo["ta_origin_source"] in {"CERT_SIA", "CERT_AIA", "ROA_AIA", "ROA_SIA"}

        # (G) הורדה מ-CROSS_TA ל-DELIVERY_TA_MISMATCH:
        # אם נראה CROSS_TA *וגם* יש CDN/hosting *וגם* TA של התוכן הגיע מרמזים תוכניים,
        # אז המסקנה: זה כנראה mismatch ברמת התעבורה (delivery), לא תלות RPKI אמיתית.
        if repo["roa_coverage_status"] == "CROSS_TA" and repo.get("is_cdn") and content_based:
            repo["delivery_ta_mismatch"] = True
            repo["delivery_ta_mismatch_reason"] = (
                f"Delivery IPs are CDN/hosting-backed ({repo.get('cdn_provider','') or repo.get('hosting_org','')}); "
                f"ROAs for delivery prefixes are under {', '.join(cov['roa_covering_tas'])}, "
                f"but repository content TA inferred from SIA/AIA is {repo.get('ta_origin','')}. "
                f"This is a transport-layer artifact (not RPKI object cross-TA dependency)."
            )
            # מסמנים סטטוס חדש ומבטלים TA_LOST כדי לא לספור כ-cross-TA אמיתי.
            repo["roa_coverage_status"] = "DELIVERY_TA_MISMATCH"
            repo["ta_lost"] = False
            repo["ta_lost_reason"] = ""
            delivery_mismatch_count += 1
        else:
            # אם זה נשאר CROSS_TA אמיתי (לא CDN mismatch), נספור כ-cross-TA.
            if repo["ta_lost"]:
                cross_ta_count += 1

    cdn_count = sum(1 for r in repositories.values() if r.get("is_cdn"))

    log(f"\n  Successfully resolved: {resolved_count} domains")
    log(f"  Failed to resolve: {failed_count} domains")
    log(f"  TRUE CROSS_TA (non-CDN delivery mismatch): {cross_ta_count} repositories")
    log(f"  Delivery TA mismatches (CDN/hosting transport): {delivery_mismatch_count} repositories")
    log(f"  Detected {cdn_count} CDN/hosting-backed repositories")

    # Phase 4 + 5 unchanged
    log("Phase 4: Parsing certificates (sampled)...")
    certificates: List[Dict[str, str]] = []
    max_certs = 50000

    for idx, cer_file in enumerate(all_cer_files[:max_certs], 1):
        if idx % 2000 == 0:
            log(f"  Parsed {idx}/{min(len(all_cer_files), max_certs)} certificates...")

        ci = parse_certificate(str(cer_file))
        if not ci:
            continue

        rel = str(cer_file.relative_to(cache_path))
        dom = extract_domain_from_path(rel)
        ci["repository"] = dom
        ci["ta_origin"] = repositories.get(dom, {}).get("ta_origin", "") or ""
        certificates.append(ci)

    log(f"  Parsed {len(certificates)} certificates")

    log("Phase 5: Generating Excel report...")
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cross_ta_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    cdn_mismatch_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws1 = wb.active
    ws1.title = "All Repositories"
    headers1 = [
        "Domain",
        "TA Origin (Expected/Display)",
        "Primary IP",
        "Resolved IPs (sample)",
        "Hosting RIR (Allocation)",
        "Hosting ASN",
        "Hosting Organization",
        "Is CDN/Hosting",
        "CDN Provider",
        "ROA Coverage Status",
        "ROA Covering TAs (Union)",
        "ROA Covering Prefixes (sample)",
        "CROSS_TA (TA_LOST)",
        "TA_LOST Reason",
        "Delivery TA Mismatch (CDN)",
        "Delivery TA Mismatch Reason",
        "CER Count",
        "ROA Count",
        "MFT Count",
        "CRL Count",
        "Total Files",
        "Sample SIA (from cert/roa)",
    ]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment

    rrow = 2
    for dom, info in sorted(repositories.items(), key=lambda x: x[0]):
        total_files = info["cer_count"] + info["roa_count"] + info["mft_count"] + info["crl_count"]
        ips_sample = ", ".join((info.get("ips_v4") or [])[:5] + (info.get("ips_v6") or [])[:2])

        ta_disp = info.get("ta_origin", "") or ""
        if is_unknown(ta_disp):
            ta_disp = "Unknown"

        ws1.cell(row=rrow, column=1, value=dom)
        ws1.cell(row=rrow, column=2, value=ta_disp)
        ws1.cell(row=rrow, column=3, value=info.get("primary_ip", ""))
        ws1.cell(row=rrow, column=4, value=ips_sample)
        ws1.cell(row=rrow, column=5, value=info.get("hosting_rir", ""))
        ws1.cell(row=rrow, column=6, value=info.get("hosting_asn") or "")
        ws1.cell(row=rrow, column=7, value=info.get("hosting_org", ""))
        ws1.cell(row=rrow, column=8, value="Yes" if info.get("is_cdn") else "No")
        ws1.cell(row=rrow, column=9, value=info.get("cdn_provider", ""))
        ws1.cell(row=rrow, column=10, value=info.get("roa_coverage_status", ""))
        ws1.cell(row=rrow, column=11, value=info.get("roa_covering_tas", ""))
        ws1.cell(row=rrow, column=12, value=info.get("roa_covering_prefixes", ""))
        ws1.cell(row=rrow, column=13, value="Yes" if info.get("ta_lost") else "No")
        ws1.cell(row=rrow, column=14, value=info.get("ta_lost_reason", ""))
        ws1.cell(row=rrow, column=15, value="Yes" if info.get("delivery_ta_mismatch") else "No")
        ws1.cell(row=rrow, column=16, value=info.get("delivery_ta_mismatch_reason", ""))
        ws1.cell(row=rrow, column=17, value=info.get("cer_count", 0))
        ws1.cell(row=rrow, column=18, value=info.get("roa_count", 0))
        ws1.cell(row=rrow, column=19, value=info.get("mft_count", 0))
        ws1.cell(row=rrow, column=20, value=info.get("crl_count", 0))
        ws1.cell(row=rrow, column=21, value=total_files)
        ws1.cell(row=rrow, column=22, value=info.get("sample_sia", ""))

        if info.get("ta_lost"):
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=rrow, column=c).fill = cross_ta_fill
        elif info.get("delivery_ta_mismatch"):
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=rrow, column=c).fill = cdn_mismatch_fill

        rrow += 1

    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 55
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["G"].width = 30
    ws1.column_dimensions["I"].width = 25
    ws1.column_dimensions["K"].width = 25
    ws1.column_dimensions["L"].width = 70
    ws1.column_dimensions["N"].width = 30
    ws1.column_dimensions["O"].width = 26
    ws1.column_dimensions["P"].width = 90
    ws1.column_dimensions["V"].width = 75

    ws2 = wb.create_sheet("CROSS_TA Repositories")
    headers2 = [
        "Domain",
        "TA Origin (Expected/Display)",
        "Primary IP",
        "Hosting RIR (Allocation)",
        "Hosting ASN",
        "Hosting Organization",
        "ROA Covering TAs (Union)",
        "ROA Covering Prefixes (sample)",
        "TA_LOST Reason",
        "Is CDN/Hosting",
        "CDN Provider",
        "Sample SIA (from cert/roa)",
    ]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment

    rr = 2
    for dom, info in sorted(repositories.items(), key=lambda x: x[0]):
        if not info.get("ta_lost"):
            continue
        ta_disp = info.get("ta_origin", "") or ""
        if is_unknown(ta_disp):
            ta_disp = "Unknown"
        ws2.cell(row=rr, column=1, value=dom)
        ws2.cell(row=rr, column=2, value=ta_disp)
        ws2.cell(row=rr, column=3, value=info.get("primary_ip", ""))
        ws2.cell(row=rr, column=4, value=info.get("hosting_rir", ""))
        ws2.cell(row=rr, column=5, value=info.get("hosting_asn") or "")
        ws2.cell(row=rr, column=6, value=info.get("hosting_org", ""))
        ws2.cell(row=rr, column=7, value=info.get("roa_covering_tas", ""))
        ws2.cell(row=rr, column=8, value=info.get("roa_covering_prefixes", ""))
        ws2.cell(row=rr, column=9, value=info.get("ta_lost_reason", ""))
        ws2.cell(row=rr, column=10, value="Yes" if info.get("is_cdn") else "No")
        ws2.cell(row=rr, column=11, value=info.get("cdn_provider", ""))
        ws2.cell(row=rr, column=12, value=info.get("sample_sia", ""))
        rr += 1

    ws2.column_dimensions["A"].width = 38
    ws2.column_dimensions["F"].width = 30
    ws2.column_dimensions["G"].width = 25
    ws2.column_dimensions["H"].width = 70
    ws2.column_dimensions["I"].width = 35
    ws2.column_dimensions["K"].width = 25
    ws2.column_dimensions["L"].width = 75

    ws5 = wb.create_sheet("CDN Delivery Mismatch")
    headers5 = [
        "Domain",
        "Expected TA (from content hints)",
        "TA Source",
        "Primary IP",
        "Is CDN/Hosting",
        "CDN Provider",
        "ROA Covering TAs (Union)",
        "ROA Coverage Status",
        "Delivery TA Mismatch Reason",
        "Sample SIA (from cert/roa)",
    ]
    for col, h in enumerate(headers5, 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment

    rr = 2
    for dom, info in sorted(repositories.items(), key=lambda x: x[0]):
        if not info.get("delivery_ta_mismatch"):
            continue
        ta_disp = info.get("ta_origin", "") or ""
        if is_unknown(ta_disp):
            ta_disp = "Unknown"
        ws5.cell(row=rr, column=1, value=dom)
        ws5.cell(row=rr, column=2, value=ta_disp)
        ws5.cell(row=rr, column=3, value=info.get("ta_origin_source", ""))
        ws5.cell(row=rr, column=4, value=info.get("primary_ip", ""))
        ws5.cell(row=rr, column=5, value="Yes" if info.get("is_cdn") else "No")
        ws5.cell(row=rr, column=6, value=info.get("cdn_provider", ""))
        ws5.cell(row=rr, column=7, value=info.get("roa_covering_tas", ""))
        ws5.cell(row=rr, column=8, value=info.get("roa_coverage_status", ""))
        ws5.cell(row=rr, column=9, value=info.get("delivery_ta_mismatch_reason", ""))
        ws5.cell(row=rr, column=10, value=info.get("sample_sia", ""))
        rr += 1

    ws5.column_dimensions["A"].width = 38
    ws5.column_dimensions["B"].width = 24
    ws5.column_dimensions["C"].width = 14
    ws5.column_dimensions["D"].width = 18
    ws5.column_dimensions["F"].width = 26
    ws5.column_dimensions["G"].width = 26
    ws5.column_dimensions["I"].width = 95
    ws5.column_dimensions["J"].width = 75

    ws3 = wb.create_sheet("All Certificates")
    headers3 = [
        "Filename",
        "Repository",
        "TA Origin (Expected/Display)",
        "Subject",
        "Issuer",
        "Not Before",
        "Not After",
        "SIA Repository",
        "SIA Manifest",
        "AIA (CA Issuers)",
    ]
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment

    rrow = 2
    for cert in certificates:
        ws3.cell(row=rrow, column=1, value=cert.get("filename", ""))
        ws3.cell(row=rrow, column=2, value=cert.get("repository", ""))
        ta_disp = cert.get("ta_origin", "") or ""
        if is_unknown(ta_disp):
            ta_disp = "Unknown"
        ws3.cell(row=rrow, column=3, value=ta_disp)
        ws3.cell(row=rrow, column=4, value=cert.get("subject", ""))
        ws3.cell(row=rrow, column=5, value=cert.get("issuer", ""))
        ws3.cell(row=rrow, column=6, value=cert.get("not_before", ""))
        ws3.cell(row=rrow, column=7, value=cert.get("not_after", ""))
        ws3.cell(row=rrow, column=8, value=cert.get("sia_repo", ""))
        ws3.cell(row=rrow, column=9, value=cert.get("sia_mft", ""))
        ws3.cell(row=rrow, column=10, value=cert.get("aia", ""))
        rrow += 1

    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["D"].width = 40
    ws3.column_dimensions["E"].width = 40
    ws3.column_dimensions["H"].width = 65
    ws3.column_dimensions["I"].width = 65
    ws3.column_dimensions["J"].width = 65

    ws4 = wb.create_sheet("Summary")
    summary = [
        ("Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Script Version", SCRIPT_VERSION),
        ("APIs Used", "Cloudflare RPKI, Cloudflare Radar, KeyCDN Geo, RIPEstat, BGPView, ip-api.com"),
        ("Cache Directory", str(cache_dir)),
        ("Output Directory", str(output_dir)),
        ("", ""),
        ("Total unique repositories", len(repositories)),
        ("Domains resolved successfully", resolved_count),
        ("Domains failed to resolve", failed_count),
        ("TRUE CROSS_TA (non-CDN delivery mismatch)", cross_ta_count),
        ("Delivery TA mismatches (CDN/hosting transport)", delivery_mismatch_count),
        ("CDN/Hosting repositories", cdn_count),
    ]
    for row_i, (k, v) in enumerate(summary, 1):
        ws4.cell(row=row_i, column=1, value=k)
        ws4.cell(row=row_i, column=2, value=v)
    ws4.column_dimensions["A"].width = 48
    ws4.column_dimensions["B"].width = 90

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = out_path / f"rpki_analysis_{ts}.xlsx"
    wb.save(out_file)

    log("Analysis complete!")
    log(f"Excel report saved to: {out_file}")

def main() -> None:
    if len(sys.argv) < 3:
        print("Usage: python3 rpki_analyze_v264.py <cache_dir> <output_dir>")
        print("Example: python3 rpki_analyze_v264.py ~/rpki-fresh-analysis/cache ~/rpki-fresh-analysis/output")
        sys.exit(1)

    analyze_rpki_data(sys.argv[1], sys.argv[2])

if __name__ == "__main__":
    main()
